#encoding:utf-8
import requests
import os

from datetime import datetime 
import time

from bs4 import BeautifulSoup
import calendar 

import collections

import openpyxl

class DataTools():
    
    def __init__(self,userid,passwd,startDate,endDate,numPass):
        self.userid=userid
        self.passwd=passwd
        self.headers={'content-type': 'application/x-www-form-urlencoded',
                   'Host':'cnshaapppwv035',
                   'Origin':'http://cnshaapppwv035',
                   'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
                   }
        self.conn=self._getConn()
        self.numPass=numPass
        self.startDate=startDate 
        self.endDate=endDate
        self.dateStaffStatue=collections.OrderedDict()#记录staff每天的statue
        self.mydates=collections.OrderedDict()
        self.mystaffs=collections.OrderedDict()
       

    '''
           获得session
    userid:用户的UUID
    passwd:密码
    
    return session
    '''
    def _getConn(self):    
        conn = requests.session()
        txtLoginID = self.userid 
        txtLoginPass=self.passwd 
        btnSubmit='Go'
        data = {'txtLoginID':txtLoginID,
            'txtLoginPass':txtLoginPass,
            'btnSubmit':btnSubmit,
            '__VIEWSTATE':r'/wEPDwUKLTYyMzE1MTcyOA9kFgICAQ9kFgICAQ8PFgIeBFRleHQFI1BsZWFzZSBlbnRlciBMb2dpbiBJRCBhbmQgUGFzc3dvcmQuZGRkz5NdtX1MuhOYXYS46IYs64VchpzAD8VP3cyJUdBlJ8U=',
            '__VIEWSTATEGENERATOR':'154FBD09',
            '__EVENTVALIDATION':r'/wEdAAq/yqky5iJYBes70AS/xuDovcB6Ys34RSNf2zaqW+GRJvOLUAKj8R+iM7xbNxCXE4885pbWlDO2hADfoPXD/5td3iL3JWK56M90wWSo2SFeP4PVTko51ft5tHXWtcNInrR9yVlMCk1cTohICZIVZoNKwzLKcr4Gt8t2jo9vjqypCl/tJUdZAGV5ufd8/RZCFysuJwy6jEQpZ2AxF2i+Ea3GScFtXgFiwzdhfqflipQ8W279k67NiGKxpstP0+jabFE=',
            'AutoLoginID':'',
            'AutoLoginContext':'',
            'AutoLoginFlag':'M',
            'AutoADLoginName':'',
            'AutoADDomainName':'',
            'AutoLoginMode':''}

         
        conn.get('http://cnshaapppwv035/WebMasterV2/pgLogin.aspx',headers=self.headers)
        
        conn.post('http://cnshaapppwv035/WebMasterV2/pgLogin.aspx',headers=self.headers,data=data)
        
        return conn
    
    
    
    '''
                    按年月获得数据并按照pandas返回
       year--年---YYYY 
       mon--月份---MM
       return soup_table:获取的请假表
    '''
   
    def _getSoupTable(self,year,mon):
        data={'ucListYearMonth$ddlYear':year,
              'ucListYearMonth$ddlMonth':mon,
              'rbView':'RbMonthly',
              'ucListOffice$ddlOffice':'CTU',
              'ucListGroup$ddlGroup':'CTU412',
              'ucListService$ddlService':'CN-SDC410H',
              '__EVENTTARGET':'ucListYearMonth$ddlMonth',
              '__EVENTARGUMENT':'',
              '__LASTFOCUS':'',
              '__VIEWSTATE':'/wEPDwUINzYyNzUwODcPZBYCAgEPZBYeAgEPZBYKAgEPDxYCHgRUZXh0BQpFcmljIE0gU2hpZGQCAw8PFgIfAGVkZAIFDw8WAh8ABQVMZWF2ZWRkAgcPDxYCHwAFmgU8YSBocmVmPScuLi9MZWF2ZS9wZ0xlYXZlQmFsYW5jZS5hc3B4Jz48Zm9udCBjbGFzcz0nTWVudVNtYWxsJz5MZWF2ZSBCYWxhbmNlPC9mb250PjwvYT48YSBocmVmPScuLi9MZWF2ZS9wZ0xlYXZlSGlzdG9yeS5hc3B4Jz48Zm9udCBjbGFzcz0nTWVudVNtYWxsJz5MZWF2ZSBTdGF0dXM8L2ZvbnQ+PC9hPjxhIGhyZWY9Jy4uL0xlYXZlL3BnQXBwbHlMZWF2ZS5hc3B4Jz48Zm9udCBjbGFzcz0nTWVudVNtYWxsJz5BcHBseSBMZWF2ZTwvZm9udD48L2E+PGEgaHJlZj0nLi4vTGVhdmUvcGdMZWF2ZUNhbGVuZGFyLmFzcHgnPjxmb250IGNsYXNzPSdNZW51U21hbGwnPkxlYXZlIENhbGVuZGFyPC9mb250PjwvYT48YSBocmVmPScuLi9MZWF2ZS9wZ0xlYXZlT25BcHByb3ZhbC5hc3B4Jz48Zm9udCBjbGFzcz0nTWVudVNtYWxsJz5BcHByb3ZlIExlYXZlPC9mb250PjwvYT48YSBocmVmPScuLi9PVC9wZ09UU3RhdHVzLmFzcHgnPjxmb250IGNsYXNzPSdNZW51U21hbGwnPk9UIFN0YXR1czwvZm9udD48L2E+PGEgaHJlZj0nLi4vT1QvcGdBcHBseU9ULmFzcHgnPjxmb250IGNsYXNzPSdNZW51U21hbGwnPkFwcGx5IE9UPC9mb250PjwvYT48YSBocmVmPScuLi9PVC9wZ09UT25BcHByb3ZhbC5hc3B4Jz48Zm9udCBjbGFzcz0nTWVudVNtYWxsJz5BcHByb3ZlIE9UPC9mb250PjwvYT5kZAIJDw8WAh8AZWRkAgMPDxYCHwAFGUxlYXZlIENhbGVuZGFyIDogSnVuIDIwMTdkZAIFDw8WAh8ABQxZZWFyIE1vbnRoIDpkZAIHD2QWBGYPEA8WBB4MQXV0b1Bvc3RCYWNrZx4LXyFEYXRhQm91bmRnZBAVBwQyMDE4BDIwMTcEMjAxNgQyMDE1BDIwMTQEMjAxMwQyMDEyFQcEMjAxOAQyMDE3BDIwMTYEMjAxNQQyMDE0BDIwMTMEMjAxMhQrAwdnZ2dnZ2dnZGQCAg8QDxYCHwFnZGRkZAIJDw8WAh4HVmlzaWJsZWhkFgJmDxAPFgQfAWcfAmdkEBUGBDIwMTgEMjAxNwQyMDE2BDIwMTUEMjAxNAQyMDEzFQYEMjAxOAQyMDE3BDIwMTYEMjAxNQQyMDE0BDIwMTMUKwMGZ2dnZ2dnFgECAWQCDw9kFgJmDxAPFgIfAmdkEBUBB0NoZW5nZHUVAQNDVFUUKwMBZxYBZmQCEQ9kFgJmDxAPFgQfAWcfAmdkEBUcD0NUVS1TREMtQURNLUNUVRZDVFUtU0RDLUFEVi0gRGVhbHMtQ1RVHkNUVS1TREMtQVNSLUNvcmUgQXNzdXJhbmNlLUNUVR5DVFUtU0RDLUFTUi1SaXNrIEFzc3VyYW5jZS1DVFUTQ1RVLVNEQy1EQVQtTCZELUNUVRpDVFUtU0RDLURBVC1PcGVyYXRpb25zLUNUVRNDVFUtU0RDLURBVC1SJlEtQ1RVHENUVS1TREMtRE9DIC0gRSBCdXNpbmVzcy1DVFUVQ1RVLVNEQy1ET0MgLSBNJkMtQ1RVHUNUVS1TREMtRE9DIC0gVHlwaW5nIFBvb2wtQ1RVD0NUVS1TREMtRklOLUNUVRFDVFUtU0RDLUhELUhSLUNUVRJDVFUtU0RDLUhELUwmRC1DVFUSQ1RVLVNEQy1IRC1TREMtQ1RVDkNUVS1TREMtSFItQ1RVEkNUVS1TREMtSU0tQURNLUNUVRJDVFUtU0RDLUlNLUZJTi1DVFUSQ1RVLVNEQy1JTS1HVFMtQ1RVEUNUVS1TREMtSU0tSFItQ1RVEkNUVS1TREMtSU0tTCZELUNUVRJDVFUtU0RDLUlNLU9QUy1DVFURQ1RVLVNEQy1JTS1RQS1DVFUSQ1RVLVNEQy1SRVMtS00tQ1RVHkNUVS1TREMtVEFYLUFjdC4gJiBQYXlyb2xsLUNUVRpDVFUtU0RDLVRBWC1Db21wbGlhbmNlLUNUVRtDVFUtU0RDLVRBWC1UUCBEb2N1bWVudC1DVFUPQ1RVLVNEQy1URUMtQ1RVD0NUVS1TREMtVFJBLUNUVRUcBkNUVTQ2NgZDVFU0MzEGQ1RVNDExBkNUVTQxMgZDVFU0NDMGQ1RVNDQxBkNUVTQ0MgZDVFU0NDgGQ1RVNDQ3BkNUVTQ0OQZDVFU0NjEGQ1RVNDcxBkNUVTQ3MgZDVFU0NzMGQ1RVNDYyBkNUVTQ4MwZDVFU0ODIGQ1RVNDg1BkNUVTQ4NgZDVFU0ODQGQ1RVNDgxBkNUVTQ4NwZDVFU0NjMGQ1RVNDIxBkNUVTQyMwZDVFU0MjIGQ1RVNDY1BkNUVTQ2NBQrAxxnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnFgECA2QCEw9kFgJmDxAPFgQfAWcfAmdkEBUFBy0tQWxsLS0PU0RDIC0gQVNSIFNEIFJBEEFTUiBTRCBSQSBUZWFtIDEQQVNSIFNEIFJBIFRlYW0gMhBBU1IgU0QgUkEgVGVhbSAzFQUACkNOLVNEQzAxLUYKQ04tU0RDNDEwRwpDTi1TREM0MTBICkNOLVNEQzQxMEkUKwMFZ2dnZ2cWAWZkAhcPZBYCZg8QDxYEHwFnHwJnZBAVBActLUFsbC0tKEFzc2lzdGFudCAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAoRGVwdXR5IFRlYW0gTWFuYWdlciAgICAgICAgICAgICAgICAgICAgIChTZW5pb3IgQXNzaXN0YW50ICAgICAgICAgICAgICAgICAgICAgICAgFQQACjcyICAgICAgICAKNzIgICAgICAgIAo3MiAgICAgICAgFCsDBGdnZ2cWAWZkAhsPZBYCZg9kFgJmD2QWAmYPZBYEZg8PFgQeCUJhY2tDb2xvcgoAHgRfIVNCAghkZAICDw8WBB8ECgAfBQIIZGQCHQ8PFgIfAAWYAVRoaXMgcGFnZSB3aWxsIGJlIGJlc3QgcHJpbnRlZCBvdXQgaW4gbGFuZHNjYXBlIGZvcm0uPGJyPkZvciBzZXR0aW5nIHBhZ2UgbWFyZ2lucywgcGxlYXNlIGdvIHRvIEZpbGUgLT4gUGFnZSBTZXR1cCBhbmQgc2V0IGxlZnQgYW5kIHJpZ2h0IG1hcmdpbiB0byA1bW0uZGQCHw8QDxYCHwNoZGRkZAIhDxAPFgIfA2hkZGRkAiMPDxYEHglTb3J0T3JkZXJnHglTb3J0RmllbGQFCVN0YWZmTmFtZWQWAgICDw8WAh8AZWRkAiUPDxYGHwZnHwcFCVN0YWZmTmFtZR8DaGQWAgICDw8WAh8AZWRkGAEFHl9fQ29udHJvbHNSZXF1aXJlUG9zdEJhY2tLZXlfXxYBBQlSYk1vbnRobHnPbk8tPVm9W3T3a9sxuUx+c1fMmcmubN6ItNqOqFgAQQ==',
              '__VIEWSTATEGENERATOR':'DA103AD2',
              '__EVENTVALIDATION':'/wEdAD9I/BGGd8o+ahVClskBzOoDY5o2V5w2o8Y1Jy8Yb477E7DNPDpVlthFU0FnnPhNFQVqBptR0RzYi3w4CDVjGkvMCJHLsBfBKIjc3bmqXVwA9iHNFoUD9v/pgoFvG21QsLHnNnyNJx00kO2DmeS1pWVj+RUHwZYfrsirv/D/HztfCpFBiWaTkEvPfzlNBQhJBsMPQ0YXEe0yj1NmOpy8nKrQSbY93YNCnGZyv2QaoM7yahMzRZoioJ40H04QUq2BuSvS2msfceBYfNm+1aXuCTzVHkJriAOMVU6RQyW27uPQzeMwFlX8Qd0qdxA/QX7KRAPEtv/Lav2S+cSJTMPR7KJh6M/s+Z4tY/PrhgMLB237mEL96Jfh0CEam5p6/IxiwMGxOnwbgtGcB8mmx2ev+FRRaYswCu2CTJqOPsYGLcV7uoKbgScuQqmLQjCPD4b7jemFQJqEK2FtNfm66tWPzfLCo3d8t98Zyucb4PxdNGVeMty+W32tV3/gyyh2w2JR9SstvqI1djEZYVHr0kX9mCL/So0B3KcZhI9GkEfJz3ZIwj2kPPthG7ftH84Vm43lc2Tru9bFNAg8BY0+R9KEeAMpUPV6J4t8lTxVqpJHu7MF4xfnTG8XN5OcH4R53mNHfOqRp2N+mgprwCclsDjCuaZaH1WhZKb/OEfJO3iDbi9He4zPbWZGpy/Q5bn3YdwIfBLNJx5Po+DjZiOlGrRa1Dhy+231LJGuZkxfcIMmSyyZggc5LCzcZT/asDBL4TbkNeIP4zm5V3NtSz5PojJ9kkQcLe1yS1f0jCEk8+mq0Pgx4kf1iFSxTnYgfw+D3vIrkkddWaeHhQ34YYmey+i5JjXhT2tstKm8c2c7HNx5d4Tzz8RvzhwbZuEeVLukmYATMelnuzr0b53SS+SkxUKM4/ygXY/FWlOfjro+WoYaLYaAs3lDUXwD+zc9OkOJYGV5Suf6Hb5YY1jiss3TPhfq3PPzng9d1lmo2KRohDIqfhrmCfyFKgsyJEzugkf/wJgr4Iky5dRJAm99oZUnQitNctVFHy9kTwgyYEqsWok820vFXVljHlDGuS1TTZy/HNNyOaTbdOyJRvXJ2dDzKf1Y/WgcEQVpqE5EvaeSxR3UNGgVsl8WrWESlCmAzSRt2AIQDymwiti+plcCNzcsRk1uLorBpLGtjDc6yTD3Z7XFx0SUEeKxifRajJqvaB74O7uRG2xUZDbSLDAjQFpPhdg6wqWl+vqf+YzHrPbz+vxmWNvWJI7bzJhCrs4nlwf1kSE7uBAqEPSqohUHUl8xCG8l3MGHC3Ums5N2E8nGfO20Y6qQcZ6kGqvis9MkPUifzAnT3vPFhAzQwFDxGI9yILvDgVsk8IXqi9dfV2sN9haPR+AFwg=='  
              }
        headers=self.headers.copy()
        headers['Referer']='http://cnshaapppwv035/SDCLeave/Leave/pgLeaveCalendar.aspx'
        r=self.conn.post('http://cnshaapppwv035/SDCLeave/Leave/pgLeaveCalendar.aspx',headers=self.headers,data=data)
        ###r就是页面的数据################下面的解析#################
        
        return BeautifulSoup(r.text,"lxml").find('table',id="dgLeaveCalendar_dgLeaveCalendar")
    
    '''
           按每月startDate和endDate去获取数据
    startDate:开始的日期，YYYYMMDD
    endDate:结束的日期,YYYYMMDD
    datesList:没有则新建，有则进行查询
    '''    
    def _getMyDatesOneMonthByStartAndEndDay(self,startDate,endDate):
        year=startDate[:4]
        mon=startDate[4:6]
        startDate=int(startDate[6:])
        endDate=int(endDate[6:])
        soup_table=self._getSoupTable(year,mon)
        
        
        for index,col in enumerate(soup_table.find('tr').find_all('td',style='color:White;')):
            if startDate<=index+1<=endDate:
                self.mydates[year+mon+self._parseDate(index+1)]=dict(year=year,mon=mon,day=index+1,week=col.br.next,statue='N')
        
        #获取每月数据
        soup_table_values=soup_table.find('tr').find_next_siblings()
    

             
        for table_row in soup_table_values:
            soup_cols=table_row.find_all('td')
            staff_id=soup_cols[1].string.strip()
            self.mystaffs[staff_id]=dict(staffName=soup_cols[0].string.strip(),staffJobTitle=soup_cols[2].string.strip())
            
            
                
            for col_index,col in enumerate(table_row.find('td',style='width:150px;',attrs={'class':'smallRecord'}).find_next_siblings()):
                if startDate<=col_index+1<=endDate:
                    date_id=year+mon+self._parseDate(col_index+1)
                    statueValue=col.find('u').string
                    if statueValue is None:
                        statueValue='N'
                    self.dateStaffStatue[staff_id+'_'+str(date_id)]=statueValue
                    
                    
        for dateId in self.mydates.keys():
            num_tmp=0
            for staffId in self.mystaffs.keys():
                if self.dateStaffStatue[staffId+'_'+dateId] in ('A','W'):
                    num_tmp=num_tmp+1
            
            if num_tmp<self.numPass:
                self.mydates[dateId]['statue']='Y'
                    
    def _parseDate(self,dateNumber):
        if dateNumber<10:
            dateNumber='0'+str(dateNumber)
        return str(dateNumber)
            
   
    def getAllData(self):
        startDate=self.startDate
        endDate=self.endDate
        monSub=abs(int(endDate[4:6])-int(startDate[4:6]))
        

        
        if monSub in [0,1,11] and (datetime.strptime(endDate,'%Y%m%d')-datetime.strptime(startDate,'%Y%m%d')).days<=31:
                
            if monSub==0:
                '一个月以内按月请求一次数据'
                self._getMyDatesOneMonthByStartAndEndDay(startDate,endDate)
                
            else:
                '两个月的分别请求数据'
                self._getMyDatesOneMonthByStartAndEndDay(startDate, startDate[:6]+str(calendar.monthrange(int(startDate[:4]),int(startDate[4:6]))[1]))
                self._getMyDatesOneMonthByStartAndEndDay(endDate[:6]+'01', endDate)
                
        else:
            print('输入的日期必须最多相差一个月')
        
def saveObjToExcel(dates,staffs,statues):        
    wb = openpyxl.Workbook()
    sht=wb.create_sheet()
    sht.cell(row=1,column=1).value="staff Name"
    sht.cell(row=1,column=2).value="staff ID"
    sht.cell(row=1,column=3).value="Job Title"
    
    for index,day in enumerate(dates):
        sht.cell(row=1,column=4+index).value=day['myKey']+'week:'+day['week']
         
    for index,staff in enumerate(staffs):
        sht.cell(row=2+index,column=1).value=staff['staffName']
        sht.cell(row=2+index,column=2).value=staff['myKey']
        sht.cell(row=2+index,column=3).value=staff['staffJobTitle']
        forKey=len(dates)
    for index,staffStatue in enumerate(statues):
        sht.cell(row=2+index//forKey,column=4+(index+1)%forKey).value=staffStatue['myValue']
    
    myPath="static\\excelFolder\\"+str(time.time())+".xlsx"
    myReturnPath="/static/excelFolder/"+str(time.time())+".xlsx"
    savePath=os.path.join(os.path.abspath('.'),myPath)    
    
    wb.save(savePath)
    return myReturnPath

if __name__ == '__main__':
       saveObjToExcel(None,None,None)



    
    
    
    


